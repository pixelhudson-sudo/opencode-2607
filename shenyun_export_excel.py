#!/usr/bin/env python3
"""Export the latest Shen Yun showtime scan to Excel."""
from datetime import datetime
from shenyun_scraper import ShenyunScraper
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "shenyun_showtimes.xlsx")

HDR_FILL = PatternFill(start_color="1C4BB6", end_color="1C4BB6", fill_type="solid")
HDR_FONT = Font(color="FFFFFF", bold=True, size=18, name="Calibri")
BODY_FONT = Font(size=14, name="Calibri")
BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALT = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)


def parse_dt(dt_str):
    if not dt_str:
        return None
    try:
        return datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None


def fmt_date(dt_str):
    dt = parse_dt(dt_str)
    if dt is None:
        return dt_str
    return dt.strftime("%y/%m/%d")


def fmt_time(dt_str):
    dt = parse_dt(dt_str)
    if dt is None:
        return dt_str
    return dt.strftime("%I:%M %p")


def clean_addr(raw):
    return (raw or "").replace("\r", "").replace("\n", ", ")


def build():
    s = ShenyunScraper()
    scans = s.query("SELECT id FROM scan_history WHERE status='completed' ORDER BY id DESC LIMIT 1")
    if not scans:
        print("No completed scans.")
        return
    sid = scans[0]["id"]

    rows = s.query(f"""
        SELECT city, country, theater_name, theater_address,
               google_maps_url, showtime_datetime, showtime_display, city_url,
               theater_group
        FROM showtimes WHERE scan_id = {sid}
        ORDER BY country, city, showtime_datetime
    """)

    wb = Workbook()
    ws = wb.active
    ws.title = "Shen Yun Showtimes"
    ws.sheet_properties.tabColor = "1C4BB6"

    headers = ["City", "Country", "Theater", "Google Maps", "Date", "Time", "Showtime", "City URL", "Group 1-8"]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ncols = len(headers)
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{len(rows)+1}"
    ws.freeze_panes = "A2"

    for ri, r in enumerate(rows, 2):
        # A: City
        c = ws.cell(row=ri, column=1, value=r["city"])
        c.font = BODY_FONT
        c.alignment = CENTER_WRAP
        c.border = BORDER

        # B: Country
        c = ws.cell(row=ri, column=2, value=r["country"])
        c.font = BODY_FONT
        c.alignment = CENTER_WRAP
        c.border = BORDER

        # C: Theater name + address
        theater = (r["theater_name"] or "").strip()
        addr = clean_addr(r["theater_address"] or "")
        c = ws.cell(row=ri, column=3)
        c.value = f"{theater}\n{addr}" if addr else theater
        c.font = BODY_FONT
        c.alignment = CENTER_WRAP
        c.border = BORDER

        # D: Google Maps link
        url = r["google_maps_url"] or ""
        c = ws.cell(row=ri, column=4)
        if url:
            c.value = "Open in Google Maps"
            c.hyperlink = url
        c.font = BODY_FONT
        c.alignment = CENTER_WRAP
        c.border = BORDER

        # E: Date
        dt_raw = r["showtime_datetime"] or ""
        c = ws.cell(row=ri, column=5, value=fmt_date(dt_raw))
        c.font = BODY_FONT
        c.alignment = CENTER
        c.border = BORDER

        # F: Time
        c = ws.cell(row=ri, column=6, value=fmt_time(dt_raw))
        c.font = BODY_FONT
        c.alignment = CENTER
        c.border = BORDER

        # G: Showtime display
        c = ws.cell(row=ri, column=7, value=r["showtime_display"] or "")
        c.font = BODY_FONT
        c.alignment = CENTER_WRAP
        c.border = BORDER

        # H: City URL link
        page_url = r["city_url"] or ""
        c = ws.cell(row=ri, column=8)
        if page_url:
            c.value = "Buy Tickets"
            c.hyperlink = page_url
        c.font = BODY_FONT
        c.alignment = CENTER_WRAP
        c.border = BORDER

        # I: Group 1-8
        c = ws.cell(row=ri, column=9, value=r["theater_group"] or "")
        c.font = BODY_FONT
        c.alignment = CENTER
        c.border = BORDER

        # Alternating row shading
        if ri % 2 == 0:
            for col in range(1, ncols + 1):
                ws.cell(row=ri, column=col).fill = ALT

    ws.row_dimensions[1].height = 24

    # Auto-fit column widths
    col_widths = {}
    for ri in range(1, len(rows) + 2):
        for ci in range(1, ncols + 1):
            cell = ws.cell(row=ri, column=ci)
            val = str(cell.value or "")
            longest_line = max((len(line) for line in val.split("\n")), default=0)
            col_widths[ci] = max(col_widths.get(ci, 0), longest_line)
    for ci, w in col_widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = min(max(w * 1.15 + 2, 10), 60)

    wb.save(OUTPUT)
    print(f"Exported {len(rows)} showtimes with {ncols} columns → {OUTPUT}")


if __name__ == "__main__":
    build()
